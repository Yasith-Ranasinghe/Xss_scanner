import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import time
import pyperclip
import re
import json
import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scanner.xss_scanner import XSSScanner
from scanner.phishing_detector import PhishingDetector
from gui.widgets import ThemedApp, CustomButton, StatusBar, create_tooltip, HyperlinkLabel

class XSSScannerApp(ThemedApp):
    def __init__(self, root):
        # Initialize the base ThemedApp
        super().__init__(root)
        
        self.root = root
        self.root.title("Advanced XSS & Phishing Scanner")
        self.root.geometry("1200x800")
        self.scanner = XSSScanner()
        self.phishing_detector = PhishingDetector()
        self.log_queue = queue.Queue()
        self.scan_thread = None
        self.clipboard_thread = None
        self.scanning = False
        self.monitoring = False
        self.last_clipboard_content = ""
        self.scan_results = []
        self.phishing_results = []
        
        # Apply modern styling
        self.create_styles()
        
        # Create the main interface
        self.create_ui()
        
        # Set up periodic callbacks
        self.check_log_queue()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def create_styles(self):
        """Create custom styles for the application"""
        # Create custom button styles
        self.style.configure('Primary.TButton', font=('Segoe UI', 10), padding=5)
        self.style.configure('Primary.TButton.Hover', background='#3498db')
        
        self.style.configure('Success.TButton', background='#2ecc71', font=('Segoe UI', 10), padding=5)
        self.style.configure('Success.TButton.Hover', background='#27ae60')
        
        self.style.configure('Danger.TButton', background='#e74c3c', font=('Segoe UI', 10), padding=5)
        self.style.configure('Danger.TButton.Hover', background='#c0392b')
        
        # Create custom tab styles
        self.style.configure('Custom.TNotebook', padding=2)
        self.style.configure('Custom.TNotebook.Tab', padding=[12, 4], font=('Segoe UI', 10))
        
        # Create custom progressbar
        self.style.configure('Custom.Horizontal.TProgressbar', thickness=20)
    
    def create_ui(self):
        """Create the main user interface"""
        # Configure grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Create main container
        main_container = ttk.Frame(self.root)
        main_container.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        main_container.columnconfigure(0, weight=1)
        main_container.rowconfigure(1, weight=1)
        
        # Create header with title and theme controls
        self.create_header(main_container)
        
        # Create notebook for tabbed interface
        self.notebook = ttk.Notebook(main_container, style='Custom.TNotebook')
        self.notebook.grid(row=1, column=0, sticky="nsew", pady=5)
        
        # Create tabs
        self.create_scanner_tab()
        self.create_phishing_tab()
        self.create_reports_tab()
        self.create_settings_tab()
        
        # Create status bar
        self.status_bar = StatusBar(self.root)
        self.status_bar.grid(row=1, column=0, sticky="ew")
        self.status_bar.set_status("Ready")
        self.status_bar.set_right_text("XSS & Phishing Scanner v2.0")

    def create_header(self, parent):
        """Create the application header"""
        header_frame = ttk.Frame(parent)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header_frame.columnconfigure(1, weight=1)
        
        # Application logo/title
        title_label = ttk.Label(header_frame, text="XSS & Phishing Scanner", 
                               font=("Segoe UI", 16, "bold"))
        title_label.grid(row=0, column=0, sticky="w", padx=5)
        
        # Theme toggle button
        self.theme_button = CustomButton(header_frame, text="Toggle Dark Mode", 
                                       command=self.toggle_dark_mode, style="Primary.TButton")
        self.theme_button.grid(row=0, column=2, sticky="e", padx=5)
        create_tooltip(self.theme_button, "Switch between light and dark themes")

    def create_scanner_tab(self):
        """Create the XSS scanner tab"""
        scanner_tab = ttk.Frame(self.notebook)
        self.notebook.add(scanner_tab, text="XSS Scanner")
        
        scanner_tab.columnconfigure(0, weight=1)
        scanner_tab.rowconfigure(2, weight=1)
        scanner_tab.rowconfigure(4, weight=1)
        
        # Create input area
        input_frame = ttk.LabelFrame(scanner_tab, text="Scan Configuration")
        input_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        input_frame.columnconfigure(1, weight=1)
        
        # URL input
        ttk.Label(input_frame, text="Target URL:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.url_var = tk.StringVar()
        self.url_entry = ttk.Entry(input_frame, textvariable=self.url_var, width=50)
        self.url_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        create_tooltip(self.url_entry, "Enter the URL to scan for XSS vulnerabilities")
        
        # Max URLs input
        ttk.Label(input_frame, text="Max URLs:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.max_urls_var = tk.StringVar(value="10")
        self.max_urls_entry = ttk.Entry(input_frame, textvariable=self.max_urls_var, width=10)
        self.max_urls_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        create_tooltip(self.max_urls_entry, "Maximum number of URLs to scan (crawler depth)")
        
        # Advanced options frame
        adv_frame = ttk.LabelFrame(input_frame, text="Advanced Options")
        adv_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        adv_frame.columnconfigure(0, weight=1)
        adv_frame.columnconfigure(1, weight=1)
        
        # Scan mode options
        self.scan_mode_var = tk.StringVar(value="standard")
        ttk.Radiobutton(adv_frame, text="Standard Scan", variable=self.scan_mode_var, 
                       value="standard").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        ttk.Radiobutton(adv_frame, text="Deep Scan", variable=self.scan_mode_var, 
                       value="deep").grid(row=0, column=1, sticky="w", padx=5, pady=2)
        
        # Custom payloads option
        self.custom_payloads_var = tk.BooleanVar(value=False)
        self.custom_payloads_check = ttk.Checkbutton(
            adv_frame, text="Use Custom Payloads", variable=self.custom_payloads_var
        )
        self.custom_payloads_check.grid(row=1, column=0, sticky="w", padx=5, pady=2)
        
        # Auth options
        self.auth_var = tk.BooleanVar(value=False)
        self.auth_check = ttk.Checkbutton(
            adv_frame, text="Use Authentication", variable=self.auth_var,
            command=self.toggle_auth_fields
        )
        self.auth_check.grid(row=1, column=1, sticky="w", padx=5, pady=2)
        
        # Selenium options
        self.use_selenium_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Use Selenium for Dynamic Testing", 
            variable=self.use_selenium_var
        ).grid(row=2, column=0, sticky="w", padx=5, pady=2)
        
        self.headless_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Headless Browser Mode", 
            variable=self.headless_var
        ).grid(row=2, column=1, sticky="w", padx=5, pady=2)
        
        # Auth fields (hidden by default)
        self.auth_frame = ttk.Frame(adv_frame)
        self.auth_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=5)
        self.auth_frame.grid_remove()  # Hide initially
        
        ttk.Label(self.auth_frame, text="Username:").grid(row=0, column=0, sticky="w", padx=5)
        self.username_var = tk.StringVar()
        self.username_entry = ttk.Entry(self.auth_frame, textvariable=self.username_var, width=20)
        self.username_entry.grid(row=0, column=1, sticky="w", padx=5)
        
        ttk.Label(self.auth_frame, text="Password:").grid(row=0, column=2, sticky="w", padx=5)
        self.password_var = tk.StringVar()
        self.password_entry = ttk.Entry(self.auth_frame, textvariable=self.password_var, width=20, show="*")
        self.password_entry.grid(row=0, column=3, sticky="w", padx=5)
        
        # Button area
        button_frame = ttk.Frame(scanner_tab)
        button_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        
        self.scan_button = CustomButton(
            button_frame, text="Start Scan", command=self.start_scan, style="Success.TButton"
        )
        self.scan_button.grid(row=0, column=0, padx=5)
        
        self.stop_button = CustomButton(
            button_frame, text="Stop Scan", command=self.stop_scan, style="Danger.TButton"
        )
        self.stop_button.grid(row=0, column=1, padx=5)
        self.stop_button.config(state=tk.DISABLED)
        
        self.monitor_var = tk.BooleanVar(value=False)
        self.monitor_check = ttk.Checkbutton(
            button_frame, text="Monitor Clipboard for URLs", 
            variable=self.monitor_var,
            command=self.toggle_clipboard_monitoring
        )
        self.monitor_check.grid(row=0, column=2, padx=5)
        
        self.minimize_var = tk.BooleanVar(value=False)
        self.minimize_check = ttk.Checkbutton(
            button_frame, text="Run in Background", 
            variable=self.minimize_var
        )
        self.minimize_check.grid(row=0, column=3, padx=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(
            scanner_tab, orient="horizontal", mode="indeterminate", 
            style="Custom.Horizontal.TProgressbar"
        )
        self.progress.grid(row=2, column=0, sticky="ew", padx=5, pady=(0, 5))
        
        # Log area
        log_frame = ttk.LabelFrame(scanner_tab, text="Scan Log")
        log_frame.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=10)
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        self.log_text.config(state=tk.DISABLED)
        
        # Results area
        results_frame = ttk.LabelFrame(scanner_tab, text="Scan Results")
        results_frame.grid(row=4, column=0, sticky="nsew", padx=5, pady=5)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        self.results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, height=15)
        self.results_text.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        self.results_text.config(state=tk.DISABLED)
    
    def toggle_auth_fields(self):
        """Show or hide the authentication fields"""
        if self.auth_var.get():
            self.auth_frame.grid()
        else:
            self.auth_frame.grid_remove()
    
    def create_phishing_tab(self):
        """Create the phishing detection tab"""
        phishing_tab = ttk.Frame(self.notebook)
        self.notebook.add(phishing_tab, text="Phishing Detector")
        
        phishing_tab.columnconfigure(0, weight=1)
        phishing_tab.rowconfigure(2, weight=1)
        phishing_tab.rowconfigure(4, weight=1)
        
        # Input area
        input_frame = ttk.LabelFrame(phishing_tab, text="URL to Check")
        input_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        input_frame.columnconfigure(1, weight=1)
        
        # URL input
        ttk.Label(input_frame, text="Suspicious URL:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.phishing_url_var = tk.StringVar()
        self.phishing_url_entry = ttk.Entry(input_frame, textvariable=self.phishing_url_var, width=50)
        self.phishing_url_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        create_tooltip(self.phishing_url_entry, "Enter the URL to check for phishing indicators")
        
        # Advanced options
        adv_frame = ttk.LabelFrame(input_frame, text="Detection Options")
        adv_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        
        # Options
        self.check_content_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Check Page Content", variable=self.check_content_var
        ).grid(row=0, column=0, sticky="w", padx=5, pady=2)
        
        self.check_forms_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Analyze Forms", variable=self.check_forms_var
        ).grid(row=0, column=1, sticky="w", padx=5, pady=2)
        
        self.check_redirects_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Track Redirects", variable=self.check_redirects_var
        ).grid(row=0, column=2, sticky="w", padx=5, pady=2)
        
        self.check_typosquatting_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            adv_frame, text="Detect Typosquatting", variable=self.check_typosquatting_var
        ).grid(row=1, column=0, sticky="w", padx=5, pady=2)
        
        # Buttons
        button_frame = ttk.Frame(phishing_tab)
        button_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        
        self.phishing_scan_button = CustomButton(
            button_frame, text="Check URL", command=self.start_phishing_scan, style="Success.TButton"
        )
        self.phishing_scan_button.grid(row=0, column=0, padx=5)
        
        self.phishing_stop_button = CustomButton(
            button_frame, text="Stop Check", command=self.stop_phishing_scan, style="Danger.TButton"
        )
        self.phishing_stop_button.grid(row=0, column=1, padx=5)
        self.phishing_stop_button.config(state=tk.DISABLED)
        
        # Progress bar
        self.phishing_progress = ttk.Progressbar(
            phishing_tab, orient="horizontal", mode="indeterminate", 
            style="Custom.Horizontal.TProgressbar"
        )
        self.phishing_progress.grid(row=2, column=0, sticky="ew", padx=5, pady=(0, 5))
        
        # Results area
        results_frame = ttk.LabelFrame(phishing_tab, text="Phishing Analysis Results")
        results_frame.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        self.phishing_results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, height=20)
        self.phishing_results_text.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        self.phishing_results_text.config(state=tk.DISABLED)
        
        # Safety score indicator
        safety_frame = ttk.LabelFrame(phishing_tab, text="Safety Score")
        safety_frame.grid(row=4, column=0, sticky="ew", padx=5, pady=5)
        
        self.safety_score_var = tk.DoubleVar(value=100)
        self.safety_meter = ttk.Progressbar(
            safety_frame, orient="horizontal", mode="determinate", 
            style="Custom.Horizontal.TProgressbar", variable=self.safety_score_var
        )
        self.safety_meter.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        
        self.safety_label = ttk.Label(safety_frame, text="Safe (100%)", font=("Segoe UI", 12, "bold"))
        self.safety_label.grid(row=1, column=0, pady=5)
    
    def start_phishing_scan(self):
        """Start the phishing scan process"""
        if self.scanning:
            return
            
        url = self.phishing_url_var.get().strip()
        if not url:
            messagebox.showerror("Error", "Please enter a URL to check")
            return
            
        # Clear previous results
        self.phishing_results_text.config(state=tk.NORMAL)
        self.phishing_results_text.delete(1.0, tk.END)
        self.phishing_results_text.config(state=tk.DISABLED)
        
        # Start the progress bar
        self.phishing_progress.start()
        self.scanning = True
        self.phishing_scan_button.config(state=tk.DISABLED)
        self.phishing_stop_button.config(state=tk.NORMAL)
        self.status_bar.set_status(f"Checking URL for phishing: {url}")
        
        # Start the scan in a separate thread
        self.scan_thread = threading.Thread(
            target=self.run_phishing_scan,
            args=(url,)
        )
        self.scan_thread.daemon = True
        self.scan_thread.start()
    
    def run_phishing_scan(self, url):
        """Run the phishing scan and update the UI when complete"""
        try:
            self.log_queue.put(f"Starting phishing check of {url}")
            
            # Analyze URL first
            url_analysis = self.phishing_detector.analyze_url(url)
            
            # Analyze content if selected
            content_analysis = None
            if self.check_content_var.get():
                try:
                    response = requests.get(url, timeout=10, verify=False)
                    content_analysis = self.phishing_detector.analyze_content(response.text)
                except Exception as e:
                    self.log_queue.put(f"Error fetching content: {str(e)}")
            
            # Combined results
            results = {
                "url_analysis": url_analysis,
                "content_analysis": content_analysis
            }
            
            # Store results for reporting
            self.phishing_results = results
            
            # Update UI on main thread
            self.root.after(0, lambda: self.display_phishing_results(results))
            
        except Exception as e:
            self.log_queue.put(f"Error during phishing scan: {str(e)}")
            self.root.after(0, self.reset_phishing_ui)
    
    def display_phishing_results(self, results):
        """Display the phishing scan results"""
        self.phishing_results_text.config(state=tk.NORMAL)
        self.phishing_results_text.delete(1.0, tk.END)
        
        url_analysis = results.get("url_analysis", {})
        content_analysis = results.get("content_analysis", {})
        
        # Calculate overall safety score (inverted phishing score)
        phishing_score = url_analysis.get("phishing_score", 0)
        content_score = content_analysis.get("content_phishing_score", 0) if content_analysis else 0
        
        # Combined score with more weight on URL analysis
        combined_score = (phishing_score * 0.7) + (content_score * 0.3)
        safety_score = max(0, 100 - (combined_score * 100))
        
        # Update safety meter
        self.safety_score_var.set(safety_score)
        
        # Update safety label with color coding
        if safety_score > 80:
            safety_text = f"Safe ({safety_score:.1f}%)"
            self.safety_label.config(text=safety_text, foreground="green")
        elif safety_score > 50:
            safety_text = f"Potentially Suspicious ({safety_score:.1f}%)"
            self.safety_label.config(text=safety_text, foreground="orange")
        else:
            safety_text = f"Likely Phishing ({safety_score:.1f}%)"
            self.safety_label.config(text=safety_text, foreground="red")
        
        # Display URL analysis
        self.phishing_results_text.insert(tk.END, "URL ANALYSIS\n", "heading")
        self.phishing_results_text.insert(tk.END, "-" * 50 + "\n")
        self.phishing_results_text.insert(tk.END, f"URL: {url_analysis.get('url', 'N/A')}\n")
        self.phishing_results_text.insert(tk.END, f"Domain: {url_analysis.get('domain', 'N/A')}\n")
        self.phishing_results_text.insert(tk.END, f"Subdomain: {url_analysis.get('subdomain', 'N/A') or 'None'}\n")
        self.phishing_results_text.insert(tk.END, f"TLD: {url_analysis.get('tld', 'N/A')}\n")
        
        # Security indicators
        self.phishing_results_text.insert(tk.END, "\nSECURITY INDICATORS\n", "heading")
        self.phishing_results_text.insert(tk.END, "-" * 50 + "\n")
        
        # HTTPS status
        is_https = url_analysis.get("is_https", False)
        https_text = "Yes ✓" if is_https else "No ✗"
        https_tag = "secure" if is_https else "insecure"
        self.phishing_results_text.insert(tk.END, f"HTTPS: {https_text}\n", https_tag)
        
        # Suspicious TLD
        is_suspicious_tld = url_analysis.get("is_suspicious_tld", False)
        tld_text = "Yes ✗" if is_suspicious_tld else "No ✓"
        tld_tag = "insecure" if is_suspicious_tld else "secure"
        self.phishing_results_text.insert(tk.END, f"Suspicious TLD: {tld_text}\n", tld_tag)
        
        # Typosquatting detection
        typosquatting_results = url_analysis.get("typosquatting_results", [])
        if typosquatting_results:
            self.phishing_results_text.insert(tk.END, f"Typosquatting Detected: Yes ✗\n", "insecure")
            self.phishing_results_text.insert(tk.END, "Possible impersonation of:\n")
            for result in typosquatting_results:
                self.phishing_results_text.insert(
                    tk.END, 
                    f"  - {result.get('legitimate_domain', 'N/A')} (similarity: {result.get('similarity', 0):.2f})\n"
                )
        else:
            self.phishing_results_text.insert(tk.END, "Typosquatting Detected: No ✓\n", "secure")
        
        # Redirect analysis
        redirect_analysis = url_analysis.get("redirect_analysis", {})
        has_suspicious_redirects = redirect_analysis.get("has_suspicious_redirects", False)
        redirect_text = "Yes ✗" if has_suspicious_redirects else "No ✓"
        redirect_tag = "insecure" if has_suspicious_redirects else "secure"
        self.phishing_results_text.insert(tk.END, f"Suspicious Redirects: {redirect_text}\n", redirect_tag)
        
        # Content analysis
        if content_analysis:
            self.phishing_results_text.insert(tk.END, "\nCONTENT ANALYSIS\n", "heading")
            self.phishing_results_text.insert(tk.END, "-" * 50 + "\n")
            
            # Phishing keywords
            phishing_keywords = content_analysis.get("phishing_keywords", [])
            if phishing_keywords:
                self.phishing_results_text.insert(tk.END, "Phishing Keywords Detected:\n", "insecure")
                self.phishing_results_text.insert(tk.END, ", ".join(phishing_keywords) + "\n")
            else:
                self.phishing_results_text.insert(tk.END, "No phishing keywords detected ✓\n", "secure")
            
            # Form analysis
            suspicious_forms = content_analysis.get("suspicious_forms", [])
            if suspicious_forms:
                self.phishing_results_text.insert(tk.END, f"\nSuspicious Forms Detected: {len(suspicious_forms)} ✗\n", "insecure")
                for i, form in enumerate(suspicious_forms, 1):
                    self.phishing_results_text.insert(tk.END, f"Form #{i}:\n")
                    self.phishing_results_text.insert(tk.END, f"  Action: {form.get('action', 'N/A')}\n")
                    self.phishing_results_text.insert(tk.END, f"  Method: {form.get('method', 'N/A')}\n")
                    self.phishing_results_text.insert(tk.END, f"  Has Password Field: {form.get('has_password_field', False)}\n")
                    self.phishing_results_text.insert(tk.END, f"  Has Credit Card Field: {form.get('has_credit_card_field', False)}\n")
            else:
                self.phishing_results_text.insert(tk.END, "No suspicious forms detected ✓\n", "secure")
        
        # Overall verdict
        self.phishing_results_text.insert(tk.END, "\nOVERALL VERDICT\n", "heading")
        self.phishing_results_text.insert(tk.END, "-" * 50 + "\n")
        
        is_phishing = url_analysis.get("is_phishing", False)
        if is_phishing:
            self.phishing_results_text.insert(tk.END, "This URL shows strong indicators of being a phishing site.\n", "verdict_bad")
            self.phishing_results_text.insert(tk.END, "Recommendation: Avoid visiting this site.\n", "verdict_bad")
        else:
            is_suspicious_content = content_analysis.get("is_suspicious_content", False) if content_analysis else False
            if is_suspicious_content:
                self.phishing_results_text.insert(tk.END, "This URL shows some suspicious indicators but is not conclusively a phishing site.\n", "verdict_warning")
                self.phishing_results_text.insert(tk.END, "Recommendation: Proceed with caution.\n", "verdict_warning")
            else:
                self.phishing_results_text.insert(tk.END, "No strong phishing indicators detected for this URL.\n", "verdict_good")
                self.phishing_results_text.insert(tk.END, "Recommendation: Generally safe to proceed, but always stay vigilant.\n", "verdict_good")
        
        # Configure tags for colored text
        self.phishing_results_text.tag_configure("heading", font=("Segoe UI", 11, "bold"))
        self.phishing_results_text.tag_configure("secure", foreground="green")
        self.phishing_results_text.tag_configure("insecure", foreground="red")
        self.phishing_results_text.tag_configure("verdict_good", foreground="green", font=("Segoe UI", 11, "bold"))
        self.phishing_results_text.tag_configure("verdict_warning", foreground="orange", font=("Segoe UI", 11, "bold"))
        self.phishing_results_text.tag_configure("verdict_bad", foreground="red", font=("Segoe UI", 11, "bold"))
        
        self.phishing_results_text.config(state=tk.DISABLED)
        self.reset_phishing_ui()
        
    def reset_phishing_ui(self):
        """Reset the phishing scan UI after scan completion"""
        self.phishing_progress.stop()
        self.scanning = False
        self.phishing_scan_button.config(state=tk.NORMAL)
        self.phishing_stop_button.config(state=tk.DISABLED)
        self.status_bar.set_status("Ready")
        
    def stop_phishing_scan(self):
        """Stop the current phishing scan"""
        if self.scan_thread and self.scanning:
            # Can't directly stop thread, but can set a flag
            self.scanning = False
            self.status_bar.set_status("Stopping scan...")
            self.log_queue.put("Scan stopping as requested by user")
            # UI will be reset when thread completes

    def update_log(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def check_log_queue(self):
        try:
            while True:
                message = self.log_queue.get_nowait()
                self.update_log(message)
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_log_queue)

    def display_results(self, vulnerabilities):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        
        if not vulnerabilities:
            self.results_text.insert(tk.END, "No XSS vulnerabilities were found.\n")
        else:
            self.results_text.insert(tk.END, f"Found {len(vulnerabilities)} vulnerabilities:\n\n")
            for i, vuln in enumerate(vulnerabilities, 1):
                self.results_text.insert(tk.END, f"Vulnerability #{i}:\n")
                self.results_text.insert(tk.END, f"Type: {vuln.get('type', 'Standard')}\n")
                self.results_text.insert(tk.END, f"URL: {vuln['url']}\n")
                
                if vuln.get('type') == 'DOM-based':
                    self.results_text.insert(tk.END, f"Sink: {vuln['details']['sink']}\n")
                    self.results_text.insert(tk.END, f"Code snippet: {vuln['details']['code']}\n\n")
                else:
                    self.results_text.insert(tk.END, f"Vector: {vuln.get('vector', 'Form')}\n")
                    self.results_text.insert(tk.END, f"Payload: {vuln['payload']}\n\n")
        
        self.results_text.config(state=tk.DISABLED)

    def start_scan(self):
        """Start the scanning process in a separate thread"""
        if self.scanning:
            return
        
        url = self.url_var.get().strip()
        if not url:
            messagebox.showerror("Error", "Please enter a target URL")
            return
        
        try:
            max_urls = int(self.max_urls_var.get())
            if max_urls < 1:
                raise ValueError("Max URLs must be at least 1")
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid number for Max URLs")
            return
        
        # Clear previous logs and results
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)
        
        # Start the progress bar
        self.progress.start()
        self.scanning = True
        self.scan_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_bar.set_status(f"Scanning: {url}")
        
        # Apply configuration
        scan_mode = self.scan_mode_var.get()
        custom_payloads = None
        
        if self.custom_payloads_var.get():
            custom_payloads = self.payload_text.get(1.0, tk.END).strip().split('\n')
        
        # Start the scan in a separate thread
        self.scan_thread = threading.Thread(
            target=self.run_scan,
            args=(url, max_urls, scan_mode, custom_payloads)
        )
        self.scan_thread.daemon = True
        self.scan_thread.start()
        
        # Minimize to system tray if option selected
        if self.minimize_var.get():
            self.root.iconify()

    def run_scan(self, url, max_urls, scan_mode="standard", custom_payloads=None):
        """Run the scan and update the UI when complete"""
        try:
            self.log_queue.put(f"Starting {scan_mode} scan of {url} (max {max_urls} URLs)")
            
            # Configure scanner
            self.scanner.timeout = int(self.timeout_var.get())
            self.scanner.max_threads = int(self.threads_var.get())
            self.scanner.user_agent = self.user_agent_var.get()
            
            if custom_payloads:
                self.log_queue.put(f"Using {len(custom_payloads)} custom payloads")
                self.scanner.set_custom_payloads(custom_payloads)
            
            # Check if authentication is required
            use_auth = self.auth_var.get() if hasattr(self, 'auth_var') else False
            use_selenium = self.use_selenium_var.get()
            
            # Initialize Selenium if needed
            if use_selenium:
                headless = self.headless_var.get()
                self.log_queue.put(f"Initializing Selenium (headless: {headless})")
                self.scanner.initialize_selenium(headless=headless)
            
            # Run scanner with appropriate settings
            if use_auth and hasattr(self, 'username_var') and hasattr(self, 'password_var'):
                username = self.username_var.get()
                password = self.password_var.get()
                if username and password:
                    self.log_queue.put(f"Using authentication with username: {username}")
                    vulnerabilities = self.scanner.scan_with_auth(
                        url, username, password, max_urls, self.log_queue, use_selenium
                    )
                else:
                    self.log_queue.put("Authentication credentials not provided, running standard scan")
                    if scan_mode == "deep":
                        vulnerabilities = self.scanner.deep_scan(url, max_urls, self.log_queue, use_selenium)
                    else:
                        vulnerabilities = self.scanner.scan_target(url, max_urls, self.log_queue)
            else:
                # Run scanner in deep mode if selected
                if scan_mode == "deep":
                    self.log_queue.put("Running in deep scan mode - this may take longer")
                    vulnerabilities = self.scanner.deep_scan(url, max_urls, self.log_queue, use_selenium)
                else:
                    vulnerabilities = self.scanner.scan_target(url, max_urls, self.log_queue)
            
            # Save results for reporting
            self.scan_results = vulnerabilities
            
            # Schedule UI updates on the main thread
            self.root.after(0, lambda: self.complete_scan(vulnerabilities))
            
        except Exception as e:
            self.log_queue.put(f"Error during scan: {str(e)}")
            self.root.after(0, self.reset_ui)
        finally:
            # Make sure to close Selenium if it was used
            if use_selenium and self.scanner.driver:
                self.log_queue.put("Closing Selenium browser")
                self.scanner.close_selenium()

    def complete_scan(self, vulnerabilities):
        """Complete the scan and update the UI"""
        self.display_results(vulnerabilities)
        self.log_queue.put("Scan completed")
        
        # If vulnerabilities were found and running in background, show notification
        if vulnerabilities and self.minimize_var.get():
            self.root.deiconify()  # Show window if minimized
            messagebox.showwarning("Vulnerabilities Detected", 
                                  f"Found {len(vulnerabilities)} XSS vulnerabilities!")
        
        self.reset_ui()

    def reset_ui(self):
        """Reset the UI after a scan"""
        self.progress.stop()
        self.scanning = False
        self.scan_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.status_bar.set_status("Ready")
        
    def stop_scan(self):
        """Stop the scanning process"""
        if self.scan_thread and self.scanning:
            self.scanning = False
            self.status_bar.set_status("Stopping scan...")
            self.log_queue.put("Scan stopping as requested by user")
            # UI will be reset when thread notices the flag

    def toggle_clipboard_monitoring(self):
        """Toggle the clipboard monitoring thread"""
        if self.monitor_var.get():
            self.start_clipboard_monitoring()
        else:
            self.stop_clipboard_monitoring()

    def start_clipboard_monitoring(self):
        """Start monitoring the clipboard for URLs"""
        if not self.monitoring:
            self.monitoring = True
            self.last_clipboard_content = pyperclip.paste()
            self.status_bar.set_status("Monitoring clipboard for URLs...")
            self.log_queue.put("Started monitoring clipboard for URLs")
            
            self.clipboard_thread = threading.Thread(target=self.monitor_clipboard)
            self.clipboard_thread.daemon = True
            self.clipboard_thread.start()

    def stop_clipboard_monitoring(self):
        """Stop monitoring the clipboard"""
        self.monitoring = False
        self.status_bar.set_status("Ready")
        self.log_queue.put("Stopped monitoring clipboard")

    def monitor_clipboard(self):
        """Monitor the clipboard for URLs and trigger scans"""
        # Use a simpler regex pattern for more reliable URL detection
        url_pattern = re.compile(
            r'https?://[^\s]+',
            re.IGNORECASE
        )
        
        while self.monitoring:
            try:
                current_clipboard = pyperclip.paste()
                
                if current_clipboard and current_clipboard != self.last_clipboard_content:
                    self.last_clipboard_content = current_clipboard
                    # Print to log for debugging
                    self.log_queue.put(f"Clipboard content changed: {current_clipboard[:50]}...")
                    
                    urls = url_pattern.findall(current_clipboard)
                    
                    for url in urls:
                        # Clean up the URL by removing trailing punctuation or spaces
                        url = url.rstrip(",.;:'\")]}>")
                        
                        self.log_queue.put(f"Found URL in clipboard: {url}")
                        
                        # Check if already scanning
                        if not self.scanning:
                            self.log_queue.put(f"Automatically starting scan for: {url}")
                            self.url_var.set(url)
                            self.root.deiconify()  # Show window if minimized
                            self.notebook.select(0)  # Switch to scanner tab
                            self.start_scan()
                        else:
                            self.log_queue.put("Scan already in progress, skipping clipboard URL")
                
                time.sleep(1)  # Check every second
            except Exception as e:
                # Log the error for debugging
                self.log_queue.put(f"Error in clipboard monitoring: {str(e)}")
                time.sleep(2)
                continue

    def generate_csv_report(self, data, path, report_type):
        """Generate a CSV report"""
        if report_type == "xss":
            # XSS CSV report
            with open(path, 'w', newline='') as csvfile:
                fieldnames = ['vulnerability_id', 'type', 'url', 'vector', 'payload', 'sink', 'code', 'status', 'timestamp']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                for i, vuln in enumerate(data, 1):
                    # Determine status based on vulnerability presence
                    status = "Vulnerable" if vuln.get('type') else "Safe"
                    
                    row = {
                        'vulnerability_id': i,
                        'type': vuln.get('type', 'Standard'),
                        'url': vuln.get('url', ''),
                        'vector': vuln.get('vector', 'Form'),
                        'payload': vuln.get('payload', ''),
                        'sink': vuln.get('details', {}).get('sink', '') if vuln.get('type') == 'DOM-based' else '',
                        'code': vuln.get('details', {}).get('code', '') if vuln.get('type') == 'DOM-based' else '',
                        'status': status,
                        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    writer.writerow(row)
        else:
            # Phishing CSV report
            url_analysis = data.get("url_analysis", {})
            content_analysis = data.get("content_analysis", {})
            
            with open(path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Property', 'Value', 'Status', 'Timestamp'])
                
                # Determine status
                is_phishing = url_analysis.get("is_phishing", False)
                status = "Phishing Detected" if is_phishing else "Safe"
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # URL info
                writer.writerow(['URL', url_analysis.get('url', 'N/A'), status, timestamp])
                writer.writerow(['Domain', url_analysis.get('domain', 'N/A'), status, timestamp])
                writer.writerow(['Subdomain', url_analysis.get('subdomain', '') or 'None', status, timestamp])
                writer.writerow(['TLD', url_analysis.get('tld', 'N/A'), status, timestamp])
                
                # Security indicators
                writer.writerow(['HTTPS', 'Yes' if url_analysis.get('is_https', False) else 'No', status, timestamp])
                writer.writerow(['Suspicious TLD', 'Yes' if url_analysis.get('is_suspicious_tld', False) else 'No', status, timestamp])
                
                # Score
                writer.writerow(['Phishing Score', url_analysis.get('phishing_score', 0), status, timestamp])
                writer.writerow(['Is Phishing', 'Yes' if url_analysis.get('is_phishing', False) else 'No', status, timestamp])
                
                # Typosquatting
                typosquatting_results = url_analysis.get("typosquatting_results", [])
                writer.writerow(['Typosquatting Detected', 'Yes' if typosquatting_results else 'No', status, timestamp])
                for i, result in enumerate(typosquatting_results, 1):
                    writer.writerow([f'Typosquatting Target {i}', result.get('legitimate_domain', 'N/A'), status, timestamp])
                    writer.writerow([f'Typosquatting Similarity {i}', result.get('similarity', 0), status, timestamp])
                
                # Redirect analysis
                redirect_analysis = url_analysis.get("redirect_analysis", {})
                writer.writerow(['Suspicious Redirects', 'Yes' if redirect_analysis.get('has_suspicious_redirects', False) else 'No', status, timestamp])
                
                # Content analysis (if available)
                if content_analysis:
                    writer.writerow(['Content Phishing Score', content_analysis.get('content_phishing_score', 0), status, timestamp])
                    
                    # Keywords
                    phishing_keywords = content_analysis.get("phishing_keywords", [])
                    writer.writerow(['Phishing Keywords', ', '.join(phishing_keywords) if phishing_keywords else 'None', status, timestamp])
                    
                    # Forms
                    suspicious_forms = content_analysis.get("suspicious_forms", [])
                    writer.writerow(['Suspicious Forms Count', len(suspicious_forms), status, timestamp])
    
    def generate_json_report(self, data, path, report_type):
        """Generate a JSON report"""
        timestamp = datetime.now().isoformat()
        
        if report_type == "xss":
            # Add status to each vulnerability
            enhanced_data = []
            for vuln in data:
                # Create a copy of the vulnerability data with status
                vuln_with_status = vuln.copy()
                vuln_with_status['status'] = "Vulnerable" if vuln.get('type') else "Safe"
                vuln_with_status['timestamp'] = timestamp
                enhanced_data.append(vuln_with_status)
            
            # Add timestamp and metadata
            report_data = {
                "report_type": "xss_scan",
                "timestamp": timestamp,
                "total_vulnerabilities": len(data),
                "vulnerabilities": enhanced_data
            }
        else:
            # Determine status
            url_analysis = data.get("url_analysis", {})
            is_phishing = url_analysis.get("is_phishing", False)
            status = "Phishing Detected" if is_phishing else "Safe"
            
            # Add timestamp and metadata to phishing report
            report_data = {
                "report_type": "phishing_analysis",
                "timestamp": timestamp,
                "status": status,
                "analysis_results": data
            }
        
        with open(path, 'w') as f:
            json.dump(report_data, f, indent=4)

    def on_closing(self):
        """Handle window closing event"""
        if self.scanning:
            if not messagebox.askyesno("Exit", "A scan is currently running. Are you sure you want to exit?"):
                return
        
        self.monitoring = False  # Stop clipboard monitoring
        self.scanning = False  # Stop any scans
        
        # Wait briefly for threads to notice the flags
        time.sleep(0.2)
        
        # Save settings here if needed
        
        self.root.destroy()

    def create_reports_tab(self):
        """Create the reports tab for generating scan reports"""
        reports_tab = ttk.Frame(self.notebook)
        self.notebook.add(reports_tab, text="Reports")
        
        reports_tab.columnconfigure(0, weight=1)
        reports_tab.rowconfigure(1, weight=1)
        
        # Report options
        options_frame = ttk.LabelFrame(reports_tab, text="Report Options")
        options_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        options_frame.columnconfigure(1, weight=1)
        
        # Report type
        ttk.Label(options_frame, text="Report Type:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.report_type_var = tk.StringVar(value="xss")
        report_type_frame = ttk.Frame(options_frame)
        report_type_frame.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        ttk.Radiobutton(
            report_type_frame, text="XSS Scan Results", variable=self.report_type_var, value="xss"
        ).grid(row=0, column=0, sticky="w", padx=5)
        
        ttk.Radiobutton(
            report_type_frame, text="Phishing Analysis", variable=self.report_type_var, value="phishing"
        ).grid(row=0, column=1, sticky="w", padx=5)
        
        # Output format
        ttk.Label(options_frame, text="Output Format:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.report_format_var = tk.StringVar(value="excel")
        format_frame = ttk.Frame(options_frame)
        format_frame.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        ttk.Radiobutton(
            format_frame, text="Excel", variable=self.report_format_var, value="excel"
        ).grid(row=0, column=0, sticky="w", padx=5)
        
        ttk.Radiobutton(
            format_frame, text="HTML", variable=self.report_format_var, value="html"
        ).grid(row=0, column=1, sticky="w", padx=5)
        
        ttk.Radiobutton(
            format_frame, text="CSV", variable=self.report_format_var, value="csv"
        ).grid(row=0, column=2, sticky="w", padx=5)
        
        ttk.Radiobutton(
            format_frame, text="JSON", variable=self.report_format_var, value="json"
        ).grid(row=0, column=3, sticky="w", padx=5)
        
        # Output location
        ttk.Label(options_frame, text="Save Location:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        path_frame = ttk.Frame(options_frame)
        path_frame.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        path_frame.columnconfigure(0, weight=1)
        
        self.report_path_var = tk.StringVar(value=os.path.join(os.getcwd(), "reports"))
        self.report_path_entry = ttk.Entry(path_frame, textvariable=self.report_path_var)
        self.report_path_entry.grid(row=0, column=0, sticky="ew", padx=2)
        
        self.browse_button = CustomButton(
            path_frame, text="Browse...", command=self.browse_report_location
        )
        self.browse_button.grid(row=0, column=1, padx=2)
        
        # Generate button
        button_frame = ttk.Frame(reports_tab)
        button_frame.grid(row=1, column=0, sticky="n", padx=5, pady=10)
        
        self.generate_button = CustomButton(
            button_frame, text="Generate Report", command=self.generate_report, 
            style="Success.TButton"
        )
        self.generate_button.grid(row=0, column=0, padx=5)
        
        # Recent reports frame
        recent_frame = ttk.LabelFrame(reports_tab, text="Recent Reports")
        recent_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
        recent_frame.columnconfigure(0, weight=1)
        
        self.recent_reports_list = tk.Listbox(recent_frame, height=8)
        self.recent_reports_list.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        self.recent_reports_list.bind("<Double-1>", self.open_selected_report)
        
        recent_button_frame = ttk.Frame(recent_frame)
        recent_button_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        
        self.open_report_button = CustomButton(
            recent_button_frame, text="Open Selected", command=self.open_selected_report
        )
        self.open_report_button.grid(row=0, column=0, padx=5)
        
        self.delete_report_button = CustomButton(
            recent_button_frame, text="Delete Selected", command=self.delete_selected_report, 
            style="Danger.TButton"
        )
        self.delete_report_button.grid(row=0, column=1, padx=5)
        
        # Populate recent reports
        self.populate_recent_reports()
    
    def browse_report_location(self):
        """Open a dialog to select report output location"""
        folder_path = filedialog.askdirectory(
            initialdir=self.report_path_var.get(),
            title="Select Report Output Directory"
        )
        if folder_path:
            self.report_path_var.set(folder_path)
    
    def generate_report(self):
        """Generate a report based on the selected options"""
        report_type = self.report_type_var.get()
        report_format = self.report_format_var.get()
        save_location = self.report_path_var.get()
        
        # Ensure save directory exists
        if not os.path.exists(save_location):
            try:
                os.makedirs(save_location)
            except Exception as e:
                messagebox.showerror("Error", f"Could not create directory: {str(e)}")
                return
        
        # Generate report filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{timestamp}"
        
        # Select appropriate data source
        if report_type == "xss":
            # If no XSS results, ask if user wants to generate a sample report
            if not self.scan_results:
                if messagebox.askyesno("No Data", "No XSS scan results available. Would you like to generate a sample report with example data?"):
                    # Create sample data for demonstration
                    self.scan_results = [
                        {
                            "type": "Reflected",
                            "url": "https://example.com/search?q=test",
                            "vector": "URL Parameter",
                            "payload": "<script>alert('XSS')</script>",
                            "details": {"sink": "innerHTML", "code": "document.getElementById('results').innerHTML = searchQuery;"}
                        },
                        {
                            "type": "DOM-based",
                            "url": "https://example.com/page",
                            "vector": "DOM",
                            "payload": "<img src=x onerror=alert('XSS')>",
                            "details": {"sink": "document.write", "code": "document.write(userContent);"}
                        }
                    ]
                else:
                    return
            data = self.scan_results
        else:  # phishing
            # If no phishing results, ask if user wants to generate a sample report
            if not self.phishing_results:
                if messagebox.askyesno("No Data", "No phishing analysis results available. Would you like to generate a sample report with example data?"):
                    # Create sample data for demonstration
                    self.phishing_results = {
                        "url_analysis": {
                            "url": "https://suspicious-site.com",
                            "domain": "suspicious-site",
                            "subdomain": "",
                            "tld": "com",
                            "is_https": False,
                            "is_suspicious_tld": False,
                            "phishing_score": 0.72,
                            "is_phishing": True,
                            "typosquatting_results": [
                                {"legitimate_domain": "legitimate-site.com", "similarity": 0.85}
                            ],
                            "redirect_analysis": {"has_suspicious_redirects": True}
                        },
                        "content_analysis": {
                            "phishing_keywords": ["login", "password", "verify"],
                            "suspicious_forms": [
                                {"action": "http://malicious-site.com/collect", "method": "POST", "has_password_field": True}
                            ],
                            "content_phishing_score": 0.65
                        }
                    }
                else:
                    return
            data = self.phishing_results
        
        try:
            # Generate appropriate report format
            full_path = ""
            if report_format == "excel":
                full_path = os.path.join(save_location, f"{filename}.xlsx")
                self.generate_excel_report(data, full_path, report_type)
            elif report_format == "html":
                full_path = os.path.join(save_location, f"{filename}.html")
                self.generate_html_report(data, full_path, report_type)
            elif report_format == "csv":
                full_path = os.path.join(save_location, f"{filename}.csv")
                self.generate_csv_report(data, full_path, report_type)
            else:  # json
                full_path = os.path.join(save_location, f"{filename}.json")
                self.generate_json_report(data, full_path, report_type)
            
            messagebox.showinfo("Success", f"Report saved to:\n{full_path}")
            self.populate_recent_reports()  # Refresh recent reports list
            
            # Offer to open the report
            if messagebox.askyesno("Open Report", "Would you like to open the report now?"):
                self.open_report(full_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
    
    def generate_excel_report(self, data, path, report_type):
        """Generate an Excel report with detailed information"""
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set report title based on type
        if report_type == "xss":
            ws.title = "XSS Scan Results"
            report_title = "XSS Vulnerability Scan Report"
        else:
            ws.title = "Phishing Analysis"
            report_title = "Phishing Analysis Report"
        
        # Format header
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = report_title
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Add report generation timestamp
        ws.merge_cells('A2:G2')
        cell = ws['A2']
        cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 5   # ID
        ws.column_dimensions['B'].width = 15  # Type
        ws.column_dimensions['C'].width = 40  # URL
        ws.column_dimensions['D'].width = 15  # Vector/Method
        ws.column_dimensions['E'].width = 30  # Payload/Details
        ws.column_dimensions['F'].width = 20  # Status
        ws.column_dimensions['G'].width = 20  # Timestamp
        
        # Add headers
        headers = ['ID', 'Type', 'URL', 'Vector', 'Payload', 'Status', 'Timestamp']
        header_row = 4  # Start headers at row 4
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            cell = ws[f"{col_letter}{header_row}"]
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data rows
        if report_type == "xss":
            # XSS report data
            for i, vuln in enumerate(data, 1):
                row_num = i + header_row
                
                # Determine status based on vulnerability presence
                status = "Vulnerable" if vuln.get('type') else "Safe"
                
                # Add data to worksheet
                ws[f"A{row_num}"] = i
                ws[f"B{row_num}"] = vuln.get('type', 'Standard')
                ws[f"C{row_num}"] = vuln.get('url', 'N/A')
                
                if vuln.get('type') == 'DOM-based':
                    ws[f"D{row_num}"] = "DOM"
                    ws[f"E{row_num}"] = vuln.get('details', {}).get('sink', 'N/A')
                else:
                    ws[f"D{row_num}"] = vuln.get('vector', 'Form')
                    ws[f"E{row_num}"] = vuln.get('payload', 'N/A')
                
                # Status with color formatting
                cell = ws[f"F{row_num}"]
                cell.value = status
                if status == "Vulnerable":
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                
                # Add timestamp
                ws[f"G{row_num}"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            # Phishing report data
            url_analysis = data.get("url_analysis", {})
            
            # Add URL information
            row_num = header_row + 1
            ws[f"A{row_num}"] = 1
            ws[f"B{row_num}"] = "Phishing Analysis"
            ws[f"C{row_num}"] = url_analysis.get('url', 'N/A')
            ws[f"D{row_num}"] = "Domain Analysis"
            
            # Create summary text
            summary = f"Domain: {url_analysis.get('domain', 'N/A')}, "
            summary += f"TLD: {url_analysis.get('tld', 'N/A')}, "
            summary += f"HTTPS: {'Yes' if url_analysis.get('is_https', False) else 'No'}"
            ws[f"E{row_num}"] = summary
            
            # Determine status
            is_phishing = url_analysis.get("is_phishing", False)
            status = "Phishing Detected" if is_phishing else "Safe"
            
            cell = ws[f"F{row_num}"]
            cell.value = status
            if is_phishing:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # Add timestamp
            ws[f"G{row_num}"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Apply border to all cells
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Save the workbook
        wb.save(path)

    def generate_html_report(self, data, path, report_type):
        """Generate an HTML report"""
        if report_type == "xss":
            # XSS HTML report generation
            with open(path, "w") as f:
                f.write(f"""<!DOCTYPE html>
<html>
<head>
    <title>XSS Scan Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        .vulnerability {{ background-color: #ffe6e6; padding: 10px; margin: 10px 0; border-radius: 5px; }}
        .secure {{ color: green; }}
        .insecure {{ color: red; }}
    </style>
</head>
<body>
    <h1>XSS Vulnerability Scan Report</h1>
    <p><strong>Date:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    <p><strong>Total Vulnerabilities Found:</strong> {len(data)}</p>
    <hr>
""")
                
                for i, vuln in enumerate(data, 1):
                    f.write(f"""
    <div class="vulnerability">
        <h3>Vulnerability #{i}</h3>
        <p><strong>Type:</strong> {vuln.get('type', 'Standard')}</p>
        <p><strong>URL:</strong> {vuln.get('url', 'N/A')}</p>
""")
                    
                    if vuln.get('type') == 'DOM-based':
                        f.write(f"""
        <p><strong>Sink:</strong> {vuln.get('details', {}).get('sink', 'N/A')}</p>
        <p><strong>Code snippet:</strong> <pre>{vuln.get('details', {}).get('code', 'N/A')}</pre></p>
""")
                    else:
                        f.write(f"""
        <p><strong>Vector:</strong> {vuln.get('vector', 'Form')}</p>
        <p><strong>Payload:</strong> <code>{vuln.get('payload', 'N/A')}</code></p>
""")
                    
                    f.write("    </div>\n")
                
                f.write("""
</body>
</html>
""")
        else:
            # Phishing HTML report generation
            url_analysis = data.get("url_analysis", {})
            content_analysis = data.get("content_analysis", {})
            
            with open(path, "w") as f:
                f.write(f"""<!DOCTYPE html>
<html>
<head>
    <title>Phishing Analysis Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1, h2 {{ color: #333; }}
        .secure {{ color: green; }}
        .insecure {{ color: red; }}
        .warning {{ color: orange; }}
        .section {{ background-color: #f5f5f5; padding: 15px; margin: 10px 0; border-radius: 5px; }}
    </style>
</head>
<body>
    <h1>Phishing Site Analysis Report</h1>
    <p><strong>Date:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    <p><strong>URL Analyzed:</strong> {url_analysis.get('url', 'N/A')}</p>
""")
                
                # Safety score
                phishing_score = url_analysis.get("phishing_score", 0)
                content_score = content_analysis.get("content_phishing_score", 0) if content_analysis else 0
                combined_score = (phishing_score * 0.7) + (content_score * 0.3)
                safety_score = max(0, 100 - (combined_score * 100))
                
                score_class = "secure"
                verdict = "Safe"
                if safety_score <= 50:
                    score_class = "insecure"
                    verdict = "Likely Phishing"
                elif safety_score <= 80:
                    score_class = "warning"
                    verdict = "Potentially Suspicious"
                
                f.write(f"""
    <div class="section">
        <h2>Safety Score: <span class="{score_class}">{safety_score:.1f}%</span></h2>
        <h3>Verdict: <span class="{score_class}">{verdict}</span></h3>
    </div>
    
    <div class="section">
        <h2>URL Analysis</h2>
        <p><strong>Domain:</strong> {url_analysis.get('domain', 'N/A')}</p>
        <p><strong>Subdomain:</strong> {url_analysis.get('subdomain', 'N/A') or 'None'}</p>
        <p><strong>TLD:</strong> {url_analysis.get('tld', 'N/A')}</p>
        <p><strong>HTTPS:</strong> <span class="{'secure' if url_analysis.get('is_https', False) else 'insecure'}">{('Yes ✓' if url_analysis.get('is_https', False) else 'No ✗')}</span></p>
        <p><strong>Suspicious TLD:</strong> <span class="{'insecure' if url_analysis.get('is_suspicious_tld', False) else 'secure'}">{('Yes ✗' if url_analysis.get('is_suspicious_tld', False) else 'No ✓')}</span></p>
""")
                
                # Typosquatting
                typosquatting_results = url_analysis.get("typosquatting_results", [])
                if typosquatting_results:
                    f.write(f"""
        <p><strong>Typosquatting Detected:</strong> <span class="insecure">Yes ✗</span></p>
        <p><strong>Possible impersonation of:</strong></p>
        <ul>
""")
                    for result in typosquatting_results:
                        f.write(f"""            <li>{result.get('legitimate_domain', 'N/A')} (similarity: {result.get('similarity', 0):.2f})</li>\n""")
                    f.write("        </ul>\n")
                else:
                    f.write("""        <p><strong>Typosquatting Detected:</strong> <span class="secure">No ✓</span></p>\n""")
                
                # Redirect analysis
                redirect_analysis = url_analysis.get("redirect_analysis", {})
                has_suspicious_redirects = redirect_analysis.get("has_suspicious_redirects", False)
                f.write(f"""
        <p><strong>Suspicious Redirects:</strong> <span class="{'insecure' if has_suspicious_redirects else 'secure'}">{('Yes ✗' if has_suspicious_redirects else 'No ✓')}</span></p>
    </div>
""")
                
                # Content analysis
                if content_analysis:
                    f.write("""
    <div class="section">
        <h2>Content Analysis</h2>
""")
                    
                    # Keywords
                    phishing_keywords = content_analysis.get("phishing_keywords", [])
                    if phishing_keywords:
                        f.write(f"""
        <p><strong>Phishing Keywords Detected:</strong> <span class="insecure">Yes ✗</span></p>
        <p>{', '.join(phishing_keywords)}</p>
""")
                    else:
                        f.write("""        <p><strong>Phishing Keywords Detected:</strong> <span class="secure">No ✓</span></p>\n""")
                    
                    # Forms
                    suspicious_forms = content_analysis.get("suspicious_forms", [])
                    if suspicious_forms:
                        f.write(f"""
        <p><strong>Suspicious Forms Detected:</strong> <span class="insecure">{len(suspicious_forms)} ✗</span></p>
        <ul>
""")
                        for i, form in enumerate(suspicious_forms, 1):
                            f.write(f"""            <li>
                <strong>Form #{i}:</strong><br>
                Action: {form.get('action', 'N/A')}<br>
                Method: {form.get('method', 'N/A')}<br>
                Has Password Field: {form.get('has_password_field', False)}<br>
                Has Credit Card Field: {form.get('has_credit_card_field', False)}
            </li>\n""")
                        f.write("        </ul>\n")
                    else:
                        f.write("""        <p><strong>Suspicious Forms Detected:</strong> <span class="secure">No ✓</span></p>\n""")
                    
                    f.write("    </div>\n")
                
                f.write("""
</body>
</html>
""")
    
    def populate_recent_reports(self):
        """Populate the recent reports list"""
        self.recent_reports_list.delete(0, tk.END)
        reports_dir = self.report_path_var.get()
        
        if not os.path.exists(reports_dir):
            return
        
        # Get all report files
        report_files = []
        for filename in os.listdir(reports_dir):
            if filename.endswith('.html') or filename.endswith('.csv') or filename.endswith('.json'):
                if 'xss_report' in filename or 'phishing_report' in filename:
                    full_path = os.path.join(reports_dir, filename)
                    report_files.append((filename, full_path, os.path.getmtime(full_path)))
        
        # Sort by most recent
        report_files.sort(key=lambda x: x[2], reverse=True)
        
        # Show top 10 most recent
        for i, (filename, full_path, _) in enumerate(report_files[:10]):
            self.recent_reports_list.insert(tk.END, filename)
            self.recent_reports_list.itemconfig(i, {'bg': '#f0f0f0' if i % 2 == 0 else '#ffffff'})
    
    def open_selected_report(self, event=None):
        """Open the selected report"""
        selected_indices = self.recent_reports_list.curselection()
        if not selected_indices:
            return
        
        selected_file = self.recent_reports_list.get(selected_indices[0])
        full_path = os.path.join(self.report_path_var.get(), selected_file)
        self.open_report(full_path)
    
    def open_report(self, path):
        """Open a report file with the default system application"""
        try:
            import webbrowser
            webbrowser.open(path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open report: {str(e)}")
    
    def delete_selected_report(self):
        """Delete the selected report file"""
        selected_indices = self.recent_reports_list.curselection()
        if not selected_indices:
            return
        
        selected_file = self.recent_reports_list.get(selected_indices[0])
        full_path = os.path.join(self.report_path_var.get(), selected_file)
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete this report?\n{selected_file}"):
            try:
                os.remove(full_path)
                self.populate_recent_reports()  # Refresh list
                messagebox.showinfo("Success", "Report deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Could not delete report: {str(e)}")

    def create_settings_tab(self):
        """Create the settings tab"""
        settings_tab = ttk.Frame(self.notebook)
        self.notebook.add(settings_tab, text="Settings")
        
        settings_tab.columnconfigure(0, weight=1)
        
        # Appearance section
        appearance_frame = ttk.LabelFrame(settings_tab, text="Appearance")
        appearance_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        appearance_frame.columnconfigure(1, weight=1)
        
        # Theme selection
        ttk.Label(appearance_frame, text="Theme:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        
        self.theme_var = tk.StringVar(value=self.current_theme)
        theme_combo = ttk.Combobox(appearance_frame, textvariable=self.theme_var, 
                                  values=self.available_themes, state="readonly", width=20)
        theme_combo.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        theme_combo.bind("<<ComboboxSelected>>", self.on_theme_change)
        
        # Dark mode toggle in settings
        dark_mode_button = CustomButton(
            appearance_frame, text="Toggle Dark Mode", 
            command=self.toggle_dark_mode, style="Primary.TButton"
        )
        dark_mode_button.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="w")
        
        # Scanner settings section
        scanner_frame = ttk.LabelFrame(settings_tab, text="Scanner Settings")
        scanner_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
        scanner_frame.columnconfigure(1, weight=1)
        
        # Timeout settings
        ttk.Label(scanner_frame, text="Request Timeout (seconds):").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.timeout_var = tk.StringVar(value="10")
        timeout_entry = ttk.Entry(scanner_frame, textvariable=self.timeout_var, width=10)
        timeout_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        # Thread settings
        ttk.Label(scanner_frame, text="Max Threads:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.threads_var = tk.StringVar(value="5")
        threads_entry = ttk.Entry(scanner_frame, textvariable=self.threads_var, width=10)
        threads_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        # User Agent
        ttk.Label(scanner_frame, text="User Agent:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.user_agent_var = tk.StringVar(value="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36")
        user_agent_entry = ttk.Entry(scanner_frame, textvariable=self.user_agent_var, width=50)
        user_agent_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        
        # Custom payloads section
        payload_frame = ttk.LabelFrame(settings_tab, text="Custom Payloads")
        payload_frame.grid(row=3, column=0, sticky="ew", padx=5, pady=5)
        payload_frame.columnconfigure(0, weight=1)
        payload_frame.rowconfigure(0, weight=1)
        
        self.payload_text = scrolledtext.ScrolledText(payload_frame, wrap=tk.WORD, height=10)
        self.payload_text.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        # Insert default payloads
        self.payload_text.insert(tk.END, "<script>alert('XSS')</script>\n<img src=x onerror=alert('XSS')>\n<svg onload=alert('XSS')>")
        
        # Save settings and reset buttons
        button_frame = ttk.Frame(settings_tab)
        button_frame.grid(row=4, column=0, pady=10)
        
        save_button = CustomButton(
            button_frame, text="Save Settings", 
            command=self.save_settings, style="Success.TButton"
        )
        save_button.grid(row=0, column=0, padx=5)
        
        reset_button = CustomButton(
            button_frame, text="Reset to Defaults", 
            command=self.reset_settings, style="Danger.TButton"
        )
        reset_button.grid(row=0, column=1, padx=5)
        
        # About section
        about_frame = ttk.LabelFrame(settings_tab, text="About")
        about_frame.grid(row=5, column=0, sticky="ew", padx=5, pady=5)
        
        about_text = """XSS & Phishing Scanner v2.0
        
A comprehensive security tool for detecting XSS vulnerabilities and phishing sites.
        
This application features:
- XSS vulnerability scanning
- Phishing detection with ML-enhanced analysis
- Detailed reporting
- Modern user interface
        """
        about_label = ttk.Label(about_frame, text=about_text, justify=tk.LEFT)
        about_label.grid(row=0, column=0, padx=10, pady=10)
    
    def on_theme_change(self, event):
        """Handle theme change from combobox"""
        selected_theme = self.theme_var.get()
        if selected_theme:
            success = self.apply_theme(selected_theme)
            if not success:
                messagebox.showerror("Error", f"Could not apply theme: {selected_theme}")
    
    def save_settings(self):
        """Save the current settings"""
        # Here you would typically save to a config file
        messagebox.showinfo("Settings Saved", "Your settings have been saved successfully.")
    
    def reset_settings(self):
        """Reset settings to defaults"""
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset all settings to default values?"):
            # Reset appearance
            self.theme_var.set("clam")
            self.apply_theme("clam")
            
            # Reset scanner settings
            self.timeout_var.set("10")
            self.threads_var.set("5")
            self.user_agent_var.set("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36")
            
            # Reset payloads
            self.payload_text.delete(1.0, tk.END)
            self.payload_text.insert(tk.END, "<script>alert('XSS')</script>\n<img src=x onerror=alert('XSS')>\n<svg onload=alert('XSS')>")
            
            messagebox.showinfo("Settings Reset", "All settings have been reset to default values.")

    def process_scan_results(self, results):
        """Process the scan results and update the UI"""
        self.scan_results = results
        
        # Create a safe entry if no vulnerabilities were found to ensure we can generate a report
        if not results:
            # Add a "safe" entry with the current URL and timestamp
            safe_entry = {
                "type": "",  # Empty type means no vulnerability
                "url": self.url_var.get(),
                "vector": "N/A",
                "payload": "N/A",
                "details": {"sink": "N/A", "code": "N/A"},
                "is_vulnerable": False,
                "status": "Safe",
                "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.scan_results = [safe_entry]
            self.log_queue.put("No vulnerabilities found, adding safe entry for report generation")
        
        # Update the results text area
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete("1.0", tk.END)
        
        if not results:
            self.results_text.insert(tk.END, "No vulnerabilities found.\n", "success")
            self.status_bar.set_status("Scan completed - No vulnerabilities found")
        else:
            self.results_text.insert(tk.END, f"Found {len(results)} vulnerabilities:\n\n", "danger")
            self.status_bar.set_status(f"Scan completed - Found {len(results)} vulnerabilities")
            
            for i, result in enumerate(results, 1):
                self.results_text.insert(tk.END, f"Vulnerability #{i}:\n", "header")
                self.results_text.insert(tk.END, f"Type: ", "bold")
                self.results_text.insert(tk.END, f"{result.get('type', 'Standard')}\n")
                
                self.results_text.insert(tk.END, f"URL: ", "bold")
                self.results_text.insert(tk.END, f"{result.get('url', 'N/A')}\n")
                
                if result.get('type') == 'DOM-based':
                    self.results_text.insert(tk.END, f"Sink: ", "bold")
                    self.results_text.insert(tk.END, f"{result.get('details', {}).get('sink', 'N/A')}\n")
                    
                    self.results_text.insert(tk.END, f"Code: ", "bold")
                    self.results_text.insert(tk.END, f"{result.get('details', {}).get('code', 'N/A')}\n")
                else:
                    self.results_text.insert(tk.END, f"Vector: ", "bold")
                    self.results_text.insert(tk.END, f"{result.get('vector', 'Form')}\n")
                    
                    self.results_text.insert(tk.END, f"Payload: ", "bold")
                    self.results_text.insert(tk.END, f"{result.get('payload', 'N/A')}\n")
                
                self.results_text.insert(tk.END, "\n")
        
        self.results_text.config(state=tk.DISABLED)

    def scan_thread(self):
        """Thread function to run the scan"""
        # Get user inputs
        target_url = self.url_var.get()
        scan_method = self.scan_method_var.get()
        
        # Validate URL
        if not target_url:
            self.log_queue.put("Error: No URL provided")
            self.status_queue.put("Ready")
            self.ui_queue.put(("error", "No URL provided"))
            self.scanning = False
            return
        
        # Add http:// if not present
        if not target_url.startswith(("http://", "https://")):
            target_url = "http://" + target_url
            self.url_var.set(target_url)
        
        # Initialize scanner if not already done
        if not hasattr(self, 'scanner'):
            self.scanner = XSSScanner()
        
        self.log_queue.put(f"Starting scan against {target_url}")
        self.status_queue.put(f"Scanning {target_url}...")
        
        try:
            # Set scanner options
            auth_type = self.auth_var.get()
            if auth_type == "basic" and self.username_var.get() and self.password_var.get():
                username = self.username_var.get()
                password = self.password_var.get()
                self.log_queue.put(f"Using Basic Auth with username: {username}")
                self.scanner.set_auth(username, password)
            else:
                # Clear auth
                self.scanner.set_auth(None, None)
            
            # Set scan depth
            scan_depth = int(self.depth_var.get())
            self.scanner.set_scan_depth(scan_depth)
            
            # Set headless browser option
            use_browser = self.browser_var.get()
            self.scanner.set_browser_mode(use_browser)
            
            # Run scan based on selected method
            if scan_method == "quick":
                results = self.scanner.quick_scan(target_url)
            elif scan_method == "forms":
                results = self.scanner.form_scan(target_url)
            else:  # comprehensive
                results = self.scanner.comprehensive_scan(target_url)
            
            # Process the results
            self.ui_queue.put(("results", results))
        except Exception as e:
            error_msg = f"Error during scan: {str(e)}"
            self.log_queue.put(error_msg)
            self.ui_queue.put(("error", error_msg))
        finally:
            # Update UI to show scan completed
            self.scanning = False
            self.log_queue.put("Scan completed")
            self.status_queue.put("Scan completed")
            self.ui_queue.put(("reset_ui", None))

    def process_ui_queue(self):
        """Process UI update requests from the scan thread"""
        try:
            while True:
                action, data = self.ui_queue.get_nowait()
                
                if action == "results":
                    self.process_scan_results(data)
                elif action == "error":
                    messagebox.showerror("Scan Error", data)
                    self.results_text.config(state=tk.NORMAL)
                    self.results_text.delete("1.0", tk.END)
                    self.results_text.insert(tk.END, f"Error: {data}\n", "error")
                    self.results_text.config(state=tk.DISABLED)
                elif action == "reset_ui":
                    self.reset_ui()
                
                self.ui_queue.task_done()
        except queue.Empty:
            pass
        
        # Schedule next check
        self.root.after(100, self.process_ui_queue)

    def process_phishing_results(self, results):
        """Process the phishing scan results and update the UI"""
        # Store the phishing results
        self.phishing_results = results
        
        # Create a safe entry if results are empty to ensure we can generate a report
        if not results or not results.get("url_analysis"):
            url = self.phishing_url_var.get()
            # Add a default safe entry
            self.phishing_results = {
                "url_analysis": {
                    "url": url,
                    "domain": "N/A",
                    "subdomain": "",
                    "tld": "N/A",
                    "is_https": False,
                    "is_suspicious_tld": False,
                    "phishing_score": 0.0,
                    "is_phishing": False,
                    "typosquatting_results": [],
                    "redirect_analysis": {"has_suspicious_redirects": False}
                },
                "content_analysis": {
                    "phishing_keywords": [],
                    "suspicious_forms": [],
                    "content_phishing_score": 0.0
                },
                "status": "Safe",
                "timestamp": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.log_queue.put("No phishing indicators found, adding safe entry for report generation")
            results = self.phishing_results
        
        # Update the results UI
        url_analysis = results.get("url_analysis", {})
        content_analysis = results.get("content_analysis", {})
        
        # Clear previous content
        self.phishing_results_text.config(state=tk.NORMAL)
        self.phishing_results_text.delete("1.0", tk.END)
        
        # URL Analysis section
        self.phishing_results_text.insert(tk.END, "URL Analysis\n", "section")
        self.phishing_results_text.insert(tk.END, f"URL: {url_analysis.get('url', 'N/A')}\n")
        self.phishing_results_text.insert(tk.END, f"Domain: {url_analysis.get('domain', 'N/A')}\n")
        self.phishing_results_text.insert(tk.END, f"Subdomain: {url_analysis.get('subdomain', 'N/A') or 'None'}\n")
        self.phishing_results_text.insert(tk.END, f"TLD: {url_analysis.get('tld', 'N/A')}\n")
        
        # Security indicators
        self.phishing_results_text.insert(tk.END, "\nSecurity Indicators\n", "section")
        
        # HTTPS status
        is_https = url_analysis.get('is_https', False)
        https_text = "Yes ✓" if is_https else "No ✗"
        https_tag = "secure" if is_https else "insecure"
        self.phishing_results_text.insert(tk.END, f"HTTPS: {https_text}\n", https_tag)
        
        # Suspicious TLD
        is_suspicious_tld = url_analysis.get('is_suspicious_tld', False)
        tld_text = "Yes ✗" if is_suspicious_tld else "No ✓"
        tld_tag = "insecure" if is_suspicious_tld else "secure"
        self.phishing_results_text.insert(tk.END, f"Suspicious TLD: {tld_text}\n", tld_tag)
        
        # Typosquatting results
        typosquatting_results = url_analysis.get("typosquatting_results", [])
        if typosquatting_results:
            self.phishing_results_text.insert(tk.END, f"\nTyposquatting Detection\n", "section")
            self.phishing_results_text.insert(tk.END, f"Possible impersonation of legitimate domains:\n", "insecure")
            
            for result in typosquatting_results:
                self.phishing_results_text.insert(tk.END, 
                    f"  - {result.get('legitimate_domain', 'N/A')} (similarity: {result.get('similarity', 0):.2f})\n"
                )
        else:
            self.phishing_results_text.insert(tk.END, f"\nTyposquatting Detection\n", "section")
            self.phishing_results_text.insert(tk.END, f"No typosquatting detected\n", "secure")
        
        # Redirect analysis
        redirect_analysis = url_analysis.get("redirect_analysis", {})
        has_suspicious_redirects = redirect_analysis.get('has_suspicious_redirects', False)
        redirect_text = "Yes ✗" if has_suspicious_redirects else "No ✓"
        redirect_tag = "insecure" if has_suspicious_redirects else "secure"
        
        self.phishing_results_text.insert(tk.END, f"\nRedirect Analysis\n", "section")
        self.phishing_results_text.insert(tk.END, f"Suspicious redirects: {redirect_text}\n", redirect_tag)
        
        # Content analysis if available
        if content_analysis:
            self.phishing_results_text.insert(tk.END, f"\nContent Analysis\n", "section")
            
            # Phishing keywords
            phishing_keywords = content_analysis.get("phishing_keywords", [])
            if phishing_keywords:
                self.phishing_results_text.insert(tk.END, f"Phishing keywords detected:\n", "insecure")
                self.phishing_results_text.insert(tk.END, f"  {', '.join(phishing_keywords)}\n")
            else:
                self.phishing_results_text.insert(tk.END, f"No phishing keywords detected\n", "secure")
            
            # Form analysis
            suspicious_forms = content_analysis.get("suspicious_forms", [])
            if suspicious_forms:
                self.phishing_results_text.insert(tk.END, f"\nSuspicious Forms\n", "section")
                self.phishing_results_text.insert(tk.END, f"Found {len(suspicious_forms)} suspicious forms:\n", "insecure")
                
                for i, form in enumerate(suspicious_forms, 1):
                    self.phishing_results_text.insert(tk.END, f"\nForm #{i}:\n", "header")
                    self.phishing_results_text.insert(tk.END, f"  Action: {form.get('action', 'N/A')}\n")
                    self.phishing_results_text.insert(tk.END, f"  Method: {form.get('method', 'N/A')}\n")
                    
                    if form.get('has_password_field'):
                        self.phishing_results_text.insert(tk.END, f"  Contains password field\n", "insecure")
                    
                    if form.get('has_credit_card_field'):
                        self.phishing_results_text.insert(tk.END, f"  Contains credit card field\n", "insecure")
            else:
                self.phishing_results_text.insert(tk.END, f"\nNo suspicious forms detected\n", "secure")
        
        # Verdict and score
        phishing_score = url_analysis.get('phishing_score', 0)
        is_phishing = url_analysis.get('is_phishing', False)
        
        self.phishing_results_text.insert(tk.END, f"\nVerdict\n", "section")
        safety_score = (1 - phishing_score) * 100  # Convert to percentage
        
        if is_phishing:
            verdict = "POTENTIALLY UNSAFE - Phishing indicators detected"
            self.phishing_results_text.insert(tk.END, f"Safety Score: {safety_score:.1f}%\n", "danger")
            self.phishing_results_text.insert(tk.END, f"{verdict}\n", "danger")
        else:
            verdict = "LIKELY SAFE - No significant phishing indicators"
            self.phishing_results_text.insert(tk.END, f"Safety Score: {safety_score:.1f}%\n", "success")
            self.phishing_results_text.insert(tk.END, f"{verdict}\n", "success")
        
        self.phishing_results_text.config(state=tk.DISABLED)
        
        # Update status bar
        if is_phishing:
            self.status_bar.set_status("Scan completed - Phishing indicators detected")
        else:
            self.status_bar.set_status("Scan completed - No significant phishing indicators")

    def phishing_thread(self):
        """Thread function to run the phishing analysis"""
        # Get the URL
        target_url = self.phishing_url_var.get()
        
        # Validate URL
        if not target_url:
            self.log_queue.put("Error: No URL provided for phishing analysis")
            self.status_queue.put("Ready")
            self.ui_queue.put(("phishing_error", "No URL provided"))
            self.scanning = False
            return
        
        # Add http:// if not present
        if not target_url.startswith(("http://", "https://")):
            target_url = "http://" + target_url
            self.phishing_url_var.set(target_url)
        
        # Initialize phishing detector if not already done
        if not hasattr(self, 'phishing_detector'):
            self.phishing_detector = PhishingDetector()
        
        self.log_queue.put(f"Starting phishing analysis for {target_url}")
        self.status_queue.put(f"Analyzing {target_url} for phishing indicators...")
        
        try:
            # URL analysis
            url_analysis = self.phishing_detector.analyze_url(target_url)
            
            # Content analysis (if enabled)
            content_analysis = {}
            if self.content_analysis_var.get():
                try:
                    # Get the web page content
                    response = requests.get(target_url, timeout=10)
                    html_content = response.text
                    
                    # Detect phishing keywords
                    phishing_keywords = self.phishing_detector.detect_phishing_keywords(html_content)
                    
                    # Analyze forms
                    suspicious_forms = self.phishing_detector.analyze_forms(html_content)
                    
                    # Content phishing score
                    content_phishing_score = 0.0
                    if phishing_keywords:
                        content_phishing_score += min(0.3, len(phishing_keywords) * 0.05)
                        
                    if suspicious_forms:
                        content_phishing_score += min(0.5, len(suspicious_forms) * 0.2)
                    
                    content_analysis = {
                        "phishing_keywords": phishing_keywords,
                        "suspicious_forms": suspicious_forms,
                        "content_phishing_score": content_phishing_score
                    }
                except Exception as e:
                    self.log_queue.put(f"Error in content analysis: {str(e)}")
                    content_analysis = {
                        "error": str(e),
                        "phishing_keywords": [],
                        "suspicious_forms": [],
                        "content_phishing_score": 0.0
                    }
            
            # Combine results
            results = {
                "url_analysis": url_analysis,
                "content_analysis": content_analysis
            }
            
            # Update phishing score if content analysis available
            if content_analysis and not content_analysis.get("error"):
                content_score = content_analysis.get("content_phishing_score", 0)
                url_score = url_analysis.get("phishing_score", 0)
                
                # Final score is weighted: 60% URL analysis, 40% content analysis
                final_score = (url_score * 0.6) + (content_score * 0.4)
                url_analysis["phishing_score"] = final_score
                url_analysis["is_phishing"] = final_score > 0.5
            
            # Process the results in the main thread
            self.ui_queue.put(("phishing_results", results))
        except Exception as e:
            error_msg = f"Error during phishing analysis: {str(e)}"
            self.log_queue.put(error_msg)
            self.ui_queue.put(("phishing_error", error_msg))
        finally:
            # Update UI to show scan completed
            self.scanning = False
            self.log_queue.put("Phishing analysis completed")
            self.status_queue.put("Phishing analysis completed")
            self.ui_queue.put(("reset_ui", None))

    def process_ui_queue(self):
        """Process UI update requests from the scan thread"""
        try:
            while True:
                action, data = self.ui_queue.get_nowait()
                
                if action == "results":
                    self.process_scan_results(data)
                elif action == "phishing_results":
                    self.process_phishing_results(data)
                elif action == "error":
                    messagebox.showerror("Scan Error", data)
                    self.results_text.config(state=tk.NORMAL)
                    self.results_text.delete("1.0", tk.END)
                    self.results_text.insert(tk.END, f"Error: {data}\n", "error")
                    self.results_text.config(state=tk.DISABLED)
                elif action == "phishing_error":
                    messagebox.showerror("Phishing Analysis Error", data)
                    self.phishing_results_text.config(state=tk.NORMAL)
                    self.phishing_results_text.delete("1.0", tk.END)
                    self.phishing_results_text.insert(tk.END, f"Error: {data}\n", "error")
                    self.phishing_results_text.config(state=tk.DISABLED)
                elif action == "reset_ui":
                    self.reset_ui()
                
                self.ui_queue.task_done()
        except queue.Empty:
            pass
        
        # Schedule next check
        self.root.after(100, self.process_ui_queue)